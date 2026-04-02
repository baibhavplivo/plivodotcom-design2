#!/usr/bin/env python3

"""
Generate the local SMS coverage number-type + pricing cache from live Plivo data.

Sources:
- src/data/sms-coverage-cache.ts (country list already fetched from the live page)
- Live SMS coverage workbook
- Live CountryPricing API
"""

from __future__ import annotations

import json
import re
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
COUNTRY_CACHE_FILE = ROOT / "src/data/sms-coverage-cache.ts"
OUTPUT_FILE = ROOT / "src/data/sms-coverage-pricing-cache.ts"
WORKBOOK_URL = (
    "https://cdn.prod.website-files.com/6565737286e587567248583f/"
    "691472d0bebd52a7d8b7b895_3f1d1f9b39853e0c22991fd810e4fd64_SMS%20Coverage%20new.xlsx"
)
COUNTRY_PRICING_URL = "https://api.plivo.com/v1/Internal/CountryPricing/?country={code}"
REFERER = "https://www.plivo.com/"
CONCURRENCY = 8

DISPLAY_TYPE_BY_PHONE_SHEET = {
    "Local": "Local Numbers",
    "Mobile": "Mobile Numbers",
    "National": "National Numbers",
    "Toll-free": "Toll-free Numbers",
    "Shortcode": "Short Codes",
}


def fetch_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"Referer": REFERER})
    with urllib.request.urlopen(request, timeout=30) as response:
        return response.read()


def fetch_json(url: str) -> dict:
    request = urllib.request.Request(url, headers={"Referer": REFERER})
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.load(response)


def load_coverage_codes() -> list[str]:
    text = COUNTRY_CACHE_FILE.read_text()
    codes = re.findall(r'\n\s+code: "([A-Z]{2})",', text)
    return list(dict.fromkeys(codes))


def load_phone_sheet_rows(workbook_bytes: bytes) -> dict[str, list[dict[str, object]]]:
    temp_path = ROOT / "tmp" / "sms-coverage-build.xlsx"
    temp_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path.write_bytes(workbook_bytes)

    workbook = load_workbook(temp_path, read_only=True, data_only=True)
    sheet = workbook["phone_numbers"]
    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    index_by_header = {header: index for index, header in enumerate(headers) if header}

    rows_by_iso: dict[str, list[dict[str, object]]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        normalized_row = list(row) + [None] * (len(headers) - len(row))
        iso = normalized_row[index_by_header["ISO"]]
        if not iso:
            continue

        row_object = {
            header: normalized_row[index]
            for header, index in index_by_header.items()
        }
        rows_by_iso.setdefault(str(iso), []).append(row_object)

    return rows_by_iso


def format_sms_rate(rate: object | None) -> str:
    if rate is None:
        return "Not Supported"

    try:
        value = float(rate)
    except (TypeError, ValueError):
        return "Not Supported"

    if value == 0:
        return "Not Supported"

    return f"${value:.4f}/sms"


def get_outbound_rate(route: dict | None) -> str:
    if not route or not route.get("outbound"):
        return "Not Supported"

    outbound = route["outbound"]
    single_rate = outbound.get("rate")
    rates = outbound.get("rates")

    if single_rate is not None:
        try:
            value = float(single_rate)
            if value != 0:
                prefix = ""
                if isinstance(rates, dict):
                    values = [float(rate) for rate in rates.values() if rate is not None]
                    if len(values) > 1 and any(rate != values[0] for rate in values):
                        prefix = "Starts at "
                return f"{prefix}${value:.4f}/sms"
        except (TypeError, ValueError):
            pass

    if isinstance(rates, dict):
        values = [float(rate) for rate in rates.values() if rate is not None]
        if values:
            minimum = min(values)
            prefix = "Starts at " if any(rate != values[0] for rate in values) else ""
            return f"{prefix}${minimum:.4f}/sms"

    return "Not Supported"


def get_min_inbound_rate(phone_numbers: list[dict], accepted_types: set[str]) -> float | None:
    values: list[float] = []

    for phone_number in phone_numbers:
        capabilities = [
            str(capability).lower()
            for capability in (phone_number.get("capabilities") or [])
        ]
        status = str(phone_number.get("status") or "")
        number_type = str(phone_number.get("number_type") or "").lower()
        inbound_rate = phone_number.get("inbound_sms_rate")

        if "sms" not in capabilities:
            continue
        if status not in {"GA", "BETA"}:
            continue
        if number_type not in accepted_types:
            continue
        if inbound_rate is None:
            continue

        values.append(float(inbound_rate))

    return min(values) if values else None


def get_mobile_inbound_rate(phone_numbers: list[dict]) -> float | None:
    for phone_number in phone_numbers:
        capabilities = [
            str(capability).lower()
            for capability in (phone_number.get("capabilities") or [])
        ]
        status = str(phone_number.get("status") or "")
        number_type = str(phone_number.get("number_type") or "").lower()
        inbound_rate = phone_number.get("inbound_sms_rate")

        if "sms" not in capabilities:
            continue
        if status not in {"GA", "BETA"}:
            continue
        if number_type != "mobile":
            continue
        if inbound_rate is None:
            continue

        return float(inbound_rate)

    return None


def build_supported_number_types(
    code: str,
    phone_sheet_rows_by_iso: dict[str, list[dict[str, object]]],
) -> list[str]:
    supported_types: list[str] = []

    if code in {"US", "CA"}:
        supported_types.append("Short Codes")

    for row in phone_sheet_rows_by_iso.get(code, []):
        if str(row.get("SMS") or "").strip().lower() != "yes":
            continue

        number_type = str(row.get("Number Type") or "").strip()
        label = DISPLAY_TYPE_BY_PHONE_SHEET.get(number_type)
        if label and label not in supported_types:
            supported_types.append(label)

    return sorted(supported_types)


def build_pricing_rows(code: str, api_data: dict) -> list[dict[str, str]]:
    phone_numbers = api_data.get("phone_numbers") or []
    sms = api_data.get("sms") or {}

    local_inbound = get_min_inbound_rate(phone_numbers, {"local", "national"})
    mobile_inbound = get_mobile_inbound_rate(phone_numbers)
    shortcode_inbound = get_min_inbound_rate(phone_numbers, {"shortcode"})
    tollfree_inbound = get_min_inbound_rate(phone_numbers, {"tollfree"})

    local_outbound = get_outbound_rate(sms.get("longcode"))
    shortcode_outbound = get_outbound_rate(sms.get("shortcode"))
    tollfree_outbound = get_outbound_rate(sms.get("tollfree"))

    rows: list[dict[str, str]] = []

    if code == "IN":
        rows.append(
            {
                "type": "Domestic",
                "outbound": "$0.00187/sms",
                "inbound": format_sms_rate(local_inbound),
            }
        )
        rows.append(
            {
                "type": "International",
                "outbound": local_outbound,
                "inbound": format_sms_rate(local_inbound),
            }
        )
        return rows

    rows.append(
        {
            "type": "Local Numbers",
            "outbound": local_outbound,
            "inbound": format_sms_rate(local_inbound),
        }
    )

    if mobile_inbound is not None:
        rows.append(
            {
                "type": "Mobile Numbers",
                "outbound": local_outbound,
                "inbound": format_sms_rate(mobile_inbound),
            }
        )

    if code in {"US", "CA"}:
        rows.append(
            {
                "type": "Short Codes",
                "outbound": shortcode_outbound,
                "inbound": format_sms_rate(shortcode_inbound),
            }
        )
        rows.append(
            {
                "type": "Toll-Free Numbers",
                "outbound": tollfree_outbound,
                "inbound": format_sms_rate(tollfree_inbound),
            }
        )

    return rows


def fetch_country_entry(code: str, phone_sheet_rows_by_iso: dict[str, list[dict[str, object]]]) -> tuple[str, dict]:
    api_data = fetch_json(COUNTRY_PRICING_URL.format(code=code))
    return (
        code,
        {
            "supportedNumberTypes": build_supported_number_types(code, phone_sheet_rows_by_iso),
            "pricingRows": build_pricing_rows(code, api_data),
        },
    )


def write_output(entries: dict[str, dict], failures: list[str]) -> None:
    contents = f"""/**
 * SMS coverage pricing cache generated from the live Plivo coverage workbook and CountryPricing API.
 * Sources:
 * - {WORKBOOK_URL}
 * - https://api.plivo.com/v1/Internal/CountryPricing/
 * Generated: {date.today().isoformat()}
 * Countries: {len(entries)} | Failures: {len(failures)}
 */

export interface SmsCoveragePricingRow {{
  type: string;
  outbound: string;
  inbound: string;
}}

export interface SmsCoverageCountryPricingData {{
  supportedNumberTypes: string[];
  pricingRows: SmsCoveragePricingRow[];
}}

export const SMS_COVERAGE_PRICING_CACHE: Record<string, SmsCoverageCountryPricingData> = {json.dumps(entries, indent=2, ensure_ascii=False)};

export const SMS_COVERAGE_PRICING_FAILURES = {json.dumps(failures, indent=2)};
"""

    OUTPUT_FILE.write_text(contents)


def main() -> None:
    codes = load_coverage_codes()
    workbook_bytes = fetch_bytes(WORKBOOK_URL)
    phone_sheet_rows_by_iso = load_phone_sheet_rows(workbook_bytes)

    entries: dict[str, dict] = {}
    failures: list[str] = []

    with ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
        futures = [
            executor.submit(fetch_country_entry, code, phone_sheet_rows_by_iso)
            for code in codes
        ]

        for future in as_completed(futures):
            try:
                code, entry = future.result()
                entries[code] = entry
            except Exception as error:  # noqa: BLE001
                failures.append(str(error))

    ordered_entries = {code: entries.get(code, {"supportedNumberTypes": [], "pricingRows": []}) for code in codes}
    write_output(ordered_entries, failures)
    print(f"Wrote {OUTPUT_FILE} with {len(ordered_entries)} countries and {len(failures)} failures")


if __name__ == "__main__":
    main()
